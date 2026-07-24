#!/usr/bin/env python3
"""
Normalize the 2026 "KIRS new data" bank-return batch into import_records.jsonl,
the exact line-schema consumed by scripts/import-bank-filings.ts.

Each source workbook is a bank's statutory filing for a quarter. Layouts are
per-bank (headers at different rows, different column orders, monthly vs
quarterly grain), so every file has an explicit spec below — nothing is guessed.

recordKind:
  TRANSACTION    -> a credit/turnover row (carries money -> counts to threshold)
  ACCOUNT_OPENED -> a new-account listing (no money -> 0 inflow, searchable only)

Output record fields (all consumed by import-bank-filings.ts):
  provider, quarter, recordKind, accountName, accountNumber, bvn, tin, nin,
  phone, email, address, amount, transactionDate, customerType, rcNumber,
  currency, txntype, desc, sourceFile, flags

Every emitted record also carries payload.importBatch (set by the importer via
IMPORT_BATCH env) so a re-run is idempotent and never touches archive rows.

Usage:  python scripts/normalize-kirs-new.py            # all specs
        python scripts/normalize-kirs-new.py --safe     # only batch=='safe'
        SRC_DIR=... OUT=... python scripts/normalize-kirs-new.py
"""
import openpyxl, os, sys, json, re

SRC = os.environ.get("SRC_DIR", r"C:\Users\olaak\Desktop\KIRS new data")
OUT = os.environ.get("OUT", r"C:\Users\olaak\AppData\Local\Temp\claude\c--laragon-www-DigitalAura\c8b834a8-1468-47f5-bf86-6a7b963d5880\scratchpad\import_records.jsonl")
ONLY_SAFE = "--safe" in sys.argv

# ------------------------------------------------------------------ helpers
def s(v):
    if v is None: return ""
    return re.sub(r"\s+", " ", str(v)).strip()

PLACEHOLDER = {"", "NIL", "NA", "N/A", "0", "-", "--", "TIN", "NIN", "BVN", "NULL", "NONE"}

def _degenerate(t):
    """True for all-zero / all-dash placeholders (e.g. 00000-0000, 0)."""
    d = re.sub(r"\D", "", t)
    return d == "" or set(d) == {"0"}

def ident(v):
    """Clean a generic id (TIN etc). Null out placeholders + all-zero fillers."""
    t = s(v)
    if t.upper() in PLACEHOLDER or _degenerate(t):
        return ""
    return t

def digits11(v):
    """BVN/NIN must be exactly 11 digits (archive precedent); else null. All-zero rejected."""
    t = re.sub(r"\D", "", s(v))
    return t if len(t) == 11 and set(t) != {"0"} else ""

def phone(v):
    t = s(v).lstrip("'").strip()
    return t

def amt(v):
    if v is None: return 0.0
    try:
        return float(re.sub(r"[,\s]", "", str(v)).replace("'", ""))
    except Exception:
        return 0.0

def ctype(v, default):
    t = s(v).upper()
    if t in ("C", "CORP", "CORPORATE"): return "CORPORATE"
    if t in ("I", "IND", "INDIVIDUAL"): return "INDIVIDUAL"
    return default

# Strong, low-false-positive corporate markers. Used only where the source has
# no reliable customer-type column (Access Q1, First Bank, UBA), so a business
# is not mis-scored against the ₦50m individual threshold instead of ₦250m.
CORP_RE = re.compile(
    r"\b(LTD|LIMITED|PLC|LLC|ENTERPRISE|ENTERPRISES|VENTURES?|COMPANY|"
    r"INTERNATIONAL|GLOBAL|SERVICES|INVESTMENTS?|RESOURCES|MULTIPURPOSE|"
    r"GOVERNMENT|ASSOCIATION|COOPERATIVE|MERCHANTS?|TRADING|CONCEPTS?|"
    r"AGENCY|COMMISSION|MINISTRY|UNIVERSITY|SCHOOL|COLLEGE|HOSPITAL|"
    r"PHARMAC\w*|NIG\b|NIGERIA)\b", re.I)

def guess_ctype(name, default):
    return "CORPORATE" if CORP_RE.search(name or "") else default

# ------------------------------------------------------------------ specs
# col indices are 0-based into the row tuple. header_row/data_row are 0-based.
# amount: either "amt" (single col) or "amt_sum" (list of cols summed).
SPECS = [
    # ---- Group A: brand-new quarters -------------------------------------
    dict(batch="safe", file="ACCESS BANK CORPORATE Q2.xlsx", sheet="Detail1",
         provider="ACCESS BANK LTD", quarter="2026-Q2", kind="TRANSACTION",
         header=2, data=3, name=0, account=1, amt=2, tin=6, email=7, phone=8, address=9,
         ctype_col=4, ctype_default="CORPORATE"),
    dict(batch="safe", file="ACCESS BANK HNW Q2 KANO.xlsx", sheet="Sheet1",
         provider="ACCESS BANK LTD", quarter="2026-Q2", kind="TRANSACTION",
         header=0, data=1, name=0, account=1, amt=2, tin=7, email=8, phone=9, address=10,
         ctype_col=4, ctype_default="INDIVIDUAL"),
    dict(batch="safe", file="FIRST BANK QUARTERLY RETURNS JAN-MARCH 2026_Updated list.xlsx", sheet="Sheet1",
         provider="FIRST BANK NIG PLC", quarter="2026-Q1", kind="TRANSACTION",
         header=2, data=3, name=1, account=2, amt=7, phone=8, nin=10, tin=11, email=13,
         ctype_default="INDIVIDUAL", guess_ctype=True),
    dict(batch="safe", file="KANO State Corporate Quarterly Returns (Transaction)- APRIL-JUNE 2026 (Q2).xlsx", sheet="Kano Corporate ",
         provider="Fidelity Bank", quarter="2026-Q2", kind="TRANSACTION",
         header=5, data=6, name=3, account=2, amt=8, tin=5, email=6, phone=7, address=4, txnmonth=1,
         ctype_default="CORPORATE"),
    dict(batch="safe", file="KANO State Individual Quarterly Returns (Transaction)- APRIL-JUNE 2026 (Q2).xlsx", sheet="Kano Corporate ",
         provider="Fidelity Bank", quarter="2026-Q2", kind="TRANSACTION",
         header=5, data=6, name=3, account=2, amt=8, nin=5, email=6, phone=7, address=4, txnmonth=1,
         ctype_default="INDIVIDUAL"),
    dict(batch="safe", file="Kano State Corporate Newly Opened Accounts- APRIL-JUNE  2026 (Q2).xlsx", sheet="KANO",
         provider="Fidelity Bank", quarter="2026-Q2", kind="ACCOUNT_OPENED",
         header=8, data=9, name=1, account=2, bvn=6, address=7, phone=11, opendate=8,
         ctype_default="CORPORATE"),
    # ---- Group B: complementary txns (existing prod rows are all 0-inflow account-opens)
    dict(batch="safe", file="ACCESS BANK First Quarter 2026 Returns of Customers with N50,000 and Above monthly transactions in Kano State Branches..xlsx", sheet="Sheet1",
         provider="ACCESS BANK LTD", quarter="2026-Q1", kind="TRANSACTION",
         header=4, data=5, name=3, address=4, phone=5, account=6, tin=7, email=8, amt_sum=[11, 12, 13],
         ctype_default="INDIVIDUAL", guess_ctype=True),
    dict(batch="safe", file="Globus Bank HNW Q1 TRANSACTION REPORT KIRS.xlsx", sheet="INDIVIDUAL",
         provider="GLOBUS BANK LTD", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=1, address=2, nin=3, email=4, phone=5, amt=9,
         ctype_default="INDIVIDUAL"),
    dict(batch="safe", file="Globus Bank HNW Q1 TRANSACTION REPORT KIRS.xlsx", sheet="CORPORATE",
         provider="GLOBUS BANK LTD", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=1, address=2, tin=3, nin=4, email=5, phone=6, amt=10,
         ctype_default="CORPORATE"),
    dict(batch="safe", file="LOTUS BANK KANO STATE RESIDENT ACCOUNTSS.xlsx", sheet="INDIVIDUAL",
         provider="LOTUS BANK LTD", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=1, address=2, nin=3, email=4, phone=5, amt=9,
         ctype_default="INDIVIDUAL"),
    dict(batch="safe", file="LOTUS BANK KANO STATE RESIDENT ACCOUNTSS.xlsx", sheet="CORPORATE",
         provider="LOTUS BANK LTD", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=1, address=2, tin=3, email=4, phone=5, amt=9,
         ctype_default="CORPORATE"),
    dict(batch="safe", file="UBA HNW Statutory returns Q1 2026.xlsx", sheet="CUMMULATIVE",
         provider="UNITED BANK FOR AFRICA", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=0, account=1, address=2, email=3, phone=4, amt=5,
         ctype_default="INDIVIDUAL", guess_ctype=True),
    # ---- Group C/D: overlap existing data -> held for explicit decision ---
    dict(batch="deferred", file="KIRS JAIZ,HNW TRX RETURNS 1ST QTR.xlsx", sheet="KIRS_QUARTERLY JAN TO MAR2026",
         provider="JAIZ BANK PLC", quarter="2026-Q1", kind="TRANSACTION",
         header=1, data=2, name=1, address=2, tin=3, nin=4, email=5, phone=6, amt=7,
         ctype_default="INDIVIDUAL"),
    dict(batch="deferred", file="KANO IRS CUMULATIVE REQUEST JAN TO MAR 2026 WEMA BANK UPDATED.xlsx", sheet="Sheet1",
         provider="WEMA BANK PLC", quarter="2026-Q1", kind="TRANSACTION",
         header=0, data=1, name=1, address=2, tin=3, nin=4, email=5, phone=6, amt=7,
         ctype_default="INDIVIDUAL"),
    dict(batch="deferred", file="JAIZKano 2 Acts opened April to June 2026.xlsx", sheet="individual",
         provider="JAIZ BANK PLC", quarter="2026-Q2", kind="ACCOUNT_OPENED",
         header=-1, data=1, name=0, account=1, address=2, ctype_default="INDIVIDUAL", dedupe_account=True),
    dict(batch="deferred", file="JAIZKano 2 Acts opened April to June 2026.xlsx", sheet="Corporate",
         provider="JAIZ BANK PLC", quarter="2026-Q2", kind="ACCOUNT_OPENED",
         header=-1, data=1, name=0, account=1, address=2, ctype_default="CORPORATE", dedupe_account=True),
    dict(batch="deferred", file="signature bank ALL ACCOUNTS OPENED - JUNE 2026 KANO BRANCH.xlsx", sheet="006 KANO",
         provider="Signature Bank Ltd", quarter="2026-Q2", kind="ACCOUNT_OPENED",
         header=6, data=7, name=2, nin=3, address=4, phone=5, email=6, ctype_default="INDIVIDUAL"),
]

def get(row, idx):
    if idx is None: return None
    return row[idx] if 0 <= idx < len(row) else None

def process(spec, out, stats):
    path = os.path.join(SRC, spec["file"])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[spec["sheet"]] if spec["sheet"] in wb.sheetnames else wb.worksheets[0]
    data_from = spec["data"]
    n = 0; total = 0.0
    seen_acct = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < data_from: continue
        name = s(get(row, spec["name"]))
        if not name: continue
        # amount
        if "amt_sum" in spec:
            a = sum(amt(get(row, c)) for c in spec["amt_sum"])
        elif "amt" in spec:
            a = amt(get(row, spec["amt"]))
        else:
            a = 0.0
        account = s(get(row, spec.get("account")))
        if spec.get("dedupe_account") and account:
            if account in seen_acct: continue
            seen_acct.add(account)
        flags = []
        if spec["kind"] == "TRANSACTION" and a <= 0:
            flags.append("zero-amount")
        rec = {
            "provider": spec["provider"],
            "quarter": spec["quarter"],
            "recordKind": spec["kind"],
            "accountName": name,
            "accountNumber": account or None,
            "bvn": digits11(get(row, spec.get("bvn"))) or None,
            "nin": digits11(get(row, spec.get("nin"))) or None,
            "tin": ident(get(row, spec.get("tin"))) or None,
            "phone": phone(get(row, spec.get("phone"))) or None,
            "email": s(get(row, spec.get("email"))) or None,
            "address": s(get(row, spec.get("address"))) or None,
            "amount": a,
            "transactionDate": s(get(row, spec.get("opendate"))) or s(get(row, spec.get("txnmonth"))) or None,
            "customerType": (guess_ctype(name, spec["ctype_default"]) if spec.get("guess_ctype")
                             else ctype(get(row, spec.get("ctype_col")), spec["ctype_default"])),
            "currency": "NGN",
            "txntype": None,
            "desc": None,
            "sourceFile": spec["file"] + "::" + spec["sheet"],
            "flags": flags,
        }
        out.write(json.dumps(rec, ensure_ascii=False) + "\n")
        n += 1; total += a
    wb.close()
    key = f'{spec["provider"]} | {spec["quarter"]} | {spec["kind"]} | {spec["batch"]}'
    stats.append((key, spec["file"] + "::" + spec["sheet"], n, total))

def main():
    specs = [sp for sp in SPECS if (not ONLY_SAFE or sp["batch"] == "safe")]
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    stats = []
    with open(OUT, "w", encoding="utf-8") as out:
        for sp in specs:
            process(sp, out, stats)
    print(f"Wrote {OUT}")
    print(f"{'PROVIDER | QUARTER | KIND | BATCH':<58} {'SOURCE':<70} {'ROWS':>8} {'SUM(NGN)':>20}")
    grand = 0; grows = 0
    for key, src, n, total in stats:
        print(f"{key:<58} {src[:68]:<70} {n:>8} {total:>20,.0f}")
        grand += total; grows += n
    print(f"{'TOTAL':<58} {'':<70} {grows:>8} {grand:>20,.0f}")

if __name__ == "__main__":
    main()
